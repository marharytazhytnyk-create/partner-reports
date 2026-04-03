"""
Bolt Food Partner Report Generator
Generates a 3-sheet Excel report:
  - Sheet 1: All active Bolt Food providers
  - Sheet 2: Easter event providers (global_cal_easter / ua-easter-2026)
  - Sheet 3: Engagement analysis by city and zone (Kyiv breakdown)
"""

import os
import re
import json
import time
import base64
import requests
import pandas as pd
from datetime import date
from io import StringIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ─── CONFIG ────────────────────────────────────────────────────────────────────
DATABRICKS_HOST  = os.getenv("DATABRICKS_HOST",  "https://bolt-incentives.cloud.databricks.com")
DATABRICKS_TOKEN = os.getenv("DATABRICKS_TOKEN", "")   # set via env var or GitHub Secret
CLUSTER_ID       = os.getenv("DATABRICKS_CLUSTER_ID", "0221-081903-9ag4bh69")

EASTER_PARTNERS_URL = (
    "https://docs.google.com/spreadsheets/d/e/"
    "2PACX-1vTEbcEMNJcF5H0tzqo4ZCQCzH_eYFKIzkvJhIAp0H_v6F2U2QzoYLk_1hlerVusfo98Z8MBJ3cryGtD"
    "/pub?gid=0&single=true&output=csv"
)

REPORT_DATE = date.today().isoformat()
OUTPUT_FILE = f"partner_report_{REPORT_DATE}.xlsx"


# ─── DATABRICKS CLUSTER API ───────────────────────────────────────────────────

def _cluster_headers() -> dict:
    return {"Authorization": f"Bearer {DATABRICKS_TOKEN}", "Content-Type": "application/json"}


def _create_context(language: str = "sql") -> str:
    resp = requests.post(
        f"{DATABRICKS_HOST}/api/1.2/contexts/create",
        headers=_cluster_headers(),
        json={"language": language, "clusterId": CLUSTER_ID},
    )
    resp.raise_for_status()
    return resp.json()["id"]


def _exec_command(ctx: str, command: str, language: str = "sql", timeout: int = 300) -> dict:
    resp = requests.post(
        f"{DATABRICKS_HOST}/api/1.2/commands/execute",
        headers=_cluster_headers(),
        json={"language": language, "clusterId": CLUSTER_ID, "contextId": ctx, "command": command},
    )
    resp.raise_for_status()
    cmd_id = resp.json()["id"]

    deadline = time.time() + timeout
    while time.time() < deadline:
        r = requests.get(
            f"{DATABRICKS_HOST}/api/1.2/commands/status",
            headers=_cluster_headers(),
            params={"clusterId": CLUSTER_ID, "contextId": ctx, "commandId": cmd_id},
        )
        r.raise_for_status()
        data = r.json()
        status = data.get("status")
        if status == "Finished":
            return data
        if status in ("Error", "Cancelled"):
            raise RuntimeError(f"Command failed: {data}")
        time.sleep(5)
    raise TimeoutError(f"Command timed out after {timeout}s")


def _rows_from_result(data: dict) -> pd.DataFrame:
    res = data.get("results", {})
    if res.get("resultType") == "error":
        raise RuntimeError(res.get("summary", "Unknown error"))
    cols = [c["name"] for c in res.get("schema", [])]
    rows = res.get("data", [])
    return pd.DataFrame(rows, columns=cols)


# ─── DATA FETCH ────────────────────────────────────────────────────────────────

def fetch_all_providers() -> pd.DataFrame:
    """Fetch all active Bolt Food providers via paginated SQL (1 000 rows per page)."""
    print("Fetching all active providers from Databricks…")

    base_sql = (
        "SELECT ptf.provider_id,"
        " ptf.provider_city_name AS city,"
        " ptf.provider_country_code AS country_code,"
        " ptf.city_zone_name AS zone,"
        " ptf.city_zone_type AS zone_type,"
        " CAST(dpp.lat AS DOUBLE) AS latitude,"
        " CAST(dpp.lng AS DOUBLE) AS longitude,"
        " CAST(COALESCE(dpp.custom_delivery_radius_in_km, 9) AS DOUBLE) AS delivery_radius_km,"
        " ptf.status, ptf.sales_segment, ptf.account_management_segment, ptf.delivery_vertical"
        " FROM ng_public_spark.etl_incentives_provider_targeting_features ptf"
        " LEFT JOIN ng_delivery_spark.delivery_provider_provider dpp ON ptf.provider_id = dpp.id"
        f" WHERE ptf.date = '{REPORT_DATE}' AND ptf.status = 'active'"
        " ORDER BY ptf.provider_country_code, ptf.provider_city_name, ptf.provider_id"
        " LIMIT {limit} OFFSET {offset}"
    )

    PAGE = 1000
    all_frames = []
    offset = 0
    ctx = _create_context("sql")

    while True:
        sql = base_sql.format(limit=PAGE, offset=offset)
        result = _exec_command(ctx, sql, "sql", timeout=300)
        df_page = _rows_from_result(result)
        if df_page.empty:
            break
        all_frames.append(df_page)
        print(f"  … page offset={offset}: {len(df_page)} rows")
        if len(df_page) < PAGE:
            break
        offset += PAGE

    df = pd.concat(all_frames, ignore_index=True) if all_frames else pd.DataFrame()
    print(f"  → {len(df):,} active providers fetched")
    return df


def fetch_easter_providers() -> pd.DataFrame:
    """Download Easter event provider list from Google Sheets (with local CSV fallback)."""
    print("Fetching Easter event providers from Google Sheets…")

    rename_map = {
        "Provider ID":                   "provider_id",
        "Brand Name":                    "brand_name",
        "City":                          "city",
        "Business Segment V2":           "business_segment",
        "Zone Name":                     "zone",
        "Account Manager Name":          "account_manager",
        "Provider Lat":                  "latitude",
        "Provider Lng":                  "longitude",
        "Delivery Radius In Km":         "delivery_radius_km",
        "Custom Delivery Radius In Km":  "custom_radius_km",
        "Total GMV Before Discounts, €": "total_gmv_eur",
    }

    def _clean(df: pd.DataFrame) -> pd.DataFrame:
        df.columns = df.columns.str.strip()
        df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})
        if "" in df.columns:
            df = df.drop(columns=[""])
        df["provider_id"] = pd.to_numeric(df["provider_id"], errors="coerce").astype("Int64")
        return df

    try:
        resp = requests.get(EASTER_PARTNERS_URL, timeout=30)
        resp.raise_for_status()
        df = _clean(pd.read_csv(StringIO(resp.text)))
        print(f"  → {len(df):,} Easter-event providers fetched")
        return df
    except Exception as exc:
        print(f"  ! Google Sheets unavailable ({exc}); falling back to local CSV")
        csv_path = os.path.join(
            os.path.dirname(__file__), "..",
            "Events Preparation Looker_02.04.csv",
        )
        df = _clean(pd.read_csv(csv_path))
        print(f"  → {len(df):,} Easter providers (from local CSV)")
        return df


# ─── EXCEL STYLING HELPERS ─────────────────────────────────────────────────────

BOLT_DARK    = "1A1A1A"
HEADER_BG    = "1DC462"
HEADER_FG    = "FFFFFF"
ALT_ROW_BG   = "F4FBF7"
BORDER_COLOR = "D0E8DB"

thin = Side(border_style="thin", color=BORDER_COLOR)


def header_style() -> dict:
    return dict(
        font=Font(bold=True, color=HEADER_FG, name="Calibri", size=11),
        fill=PatternFill("solid", fgColor=HEADER_BG),
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
        border=Border(bottom=Side("medium", color="009933")),
    )


def apply_style(cell, **kwargs):
    for attr, val in kwargs.items():
        setattr(cell, attr, val)


def write_dataframe(ws, df: pd.DataFrame, start_row: int = 4):
    """Write DataFrame to worksheet: header on start_row-1, data from start_row."""
    hs = header_style()
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row - 1, column=col_idx, value=col_name)
        apply_style(cell, **hs)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(40, len(str(col_name)) + 4))

    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        fill_color = ALT_ROW_BG if r_idx % 2 == 0 else "FFFFFF"
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = ws.cell(row=start_row, column=1)


def add_title_row(ws, title: str, subtitle: str, ncols: int):
    """Dark title on row 1, light subtitle on row 2 (if provided)."""
    ncols = max(ncols, 3)
    ws.row_dimensions[1].height = 30
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    cell.fill      = PatternFill("solid", fgColor=BOLT_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

    if subtitle:
        ws.row_dimensions[2].height = 20
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font      = Font(italic=True, size=10, color="444444", name="Calibri")
        sub.fill      = PatternFill("solid", fgColor="E8F7EE")
        sub.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)


# ─── REPORT BUILDER ────────────────────────────────────────────────────────────

def build_report(df_all: pd.DataFrame, df_easter: pd.DataFrame, output_path: str):
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Partners ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Partners"

    if not df_all.empty:
        display_cols = [c for c in [
            "provider_id", "city", "country_code", "zone", "zone_type",
            "latitude", "longitude", "delivery_radius_km",
            "status", "sales_segment", "delivery_vertical",
        ] if c in df_all.columns]
        label_map = {
            "provider_id": "Provider ID", "city": "City", "country_code": "Country",
            "zone": "Zone Name", "zone_type": "Zone Type",
            "latitude": "Latitude", "longitude": "Longitude",
            "delivery_radius_km": "Delivery Radius (km)",
            "status": "Status", "sales_segment": "Sales Segment",
            "delivery_vertical": "Vertical",
        }
        df1 = df_all[display_cols].copy()
        df1.columns = [label_map.get(c, c) for c in display_cols]
    else:
        df1 = pd.DataFrame(columns=["Provider ID", "City", "Country", "Zone Name",
                                     "Latitude", "Longitude", "Delivery Radius (km)"])

    add_title_row(ws1, f"All Active Bolt Food Partners — {REPORT_DATE}",
                  f"Total: {len(df1):,} providers", len(df1.columns))
    write_dataframe(ws1, df1, start_row=4)

    # ── Sheet 2: Easter Event Partners ────────────────────────────────────────
    ws2 = wb.create_sheet("Easter Event Partners")

    display_cols2 = [c for c in [
        "provider_id", "brand_name", "city", "zone",
        "latitude", "longitude", "delivery_radius_km", "custom_radius_km",
        "business_segment", "account_manager",
    ] if c in df_easter.columns]
    label_map2 = {
        "provider_id": "Provider ID", "brand_name": "Brand Name", "city": "City",
        "zone": "Zone Name", "latitude": "Latitude", "longitude": "Longitude",
        "delivery_radius_km": "Delivery Radius (km)", "custom_radius_km": "Custom Radius (km)",
        "business_segment": "Business Segment", "account_manager": "Account Manager",
    }
    df2 = df_easter[display_cols2].copy()
    df2.columns = [label_map2.get(c, c) for c in display_cols2]

    add_title_row(ws2, f"Easter Event Partners (global_cal_easter) — {REPORT_DATE}",
                  f"Total: {len(df2):,} providers", len(df2.columns))
    write_dataframe(ws2, df2, start_row=4)

    # ── Sheet 3: Engagement Analysis ─────────────────────────────────────────
    ws3 = wb.create_sheet("Engagement Analysis")
    ws3.sheet_view.showGridLines = False

    easter_ids = set(df_easter["provider_id"].dropna().astype(int).tolist())

    if not df_all.empty:
        df_all = df_all.copy()
        df_all["is_easter"] = df_all["provider_id"].astype(float).astype("Int64").isin(easter_ids)
        city_stats = (
            df_all.groupby("city")
            .agg(total_providers=("provider_id", "count"), easter_providers=("is_easter", "sum"))
            .reset_index()
        )
    else:
        city_stats = (
            df_easter.groupby("city")
            .agg(easter_providers=("provider_id", "count"))
            .reset_index()
        )
        city_stats["total_providers"] = city_stats["easter_providers"]

    city_stats["engagement_pct"] = (
        city_stats["easter_providers"] / city_stats["total_providers"] * 100
    ).round(1)
    city_stats = city_stats.sort_values("easter_providers", ascending=False).reset_index(drop=True)

    # Kyiv zone breakdown
    kyiv_easter = df_easter[df_easter["city"].str.lower().str.contains("kyiv|kiev|київ", na=False)]
    if not df_all.empty:
        kyiv_all = df_all[df_all["city"].str.lower().str.contains("kyiv|kiev|київ", na=False)].copy()
        zone_stats = (
            kyiv_all.groupby("zone")
            .agg(total_providers=("provider_id", "count"), easter_providers=("is_easter", "sum"))
            .reset_index()
        )
    else:
        zone_stats = (
            kyiv_easter.groupby("zone")
            .agg(easter_providers=("provider_id", "count"))
            .reset_index()
        )
        zone_stats["total_providers"] = zone_stats["easter_providers"]

    zone_stats["engagement_pct"] = (
        zone_stats["easter_providers"] / zone_stats["total_providers"] * 100
    ).round(1)
    zone_stats = zone_stats.sort_values("easter_providers", ascending=False).reset_index(drop=True)

    add_title_row(ws3, f"Partner Engagement — Easter Event — {REPORT_DATE}", "", 8)

    # KPIs
    row = 3
    total_providers_count = len(df_all) if not df_all.empty else 0
    kpi_data = [
        ("Total Active Providers",         f"{total_providers_count:,}" if total_providers_count else "N/A"),
        ("Easter Event Providers",          f"{len(df_easter):,}"),
        ("Overall Engagement Rate",
         f"{len(df_easter)/total_providers_count*100:.1f}%" if total_providers_count else "N/A"),
        ("Cities with Easter Participants", str(city_stats[city_stats["easter_providers"] > 0].shape[0])),
        ("Report Date",                     REPORT_DATE),
    ]
    ws3.cell(row=row, column=1, value="Key Metrics").font = Font(bold=True, size=12, name="Calibri")
    ws3.cell(row=row, column=1).fill = PatternFill("solid", fgColor="E8F7EE")
    row += 1
    for label, value in kpi_data:
        lc = ws3.cell(row=row, column=1, value=label)
        vc = ws3.cell(row=row, column=2, value=value)
        lc.font   = Font(bold=True, name="Calibri", size=10)
        lc.fill   = PatternFill("solid", fgColor="F0FAF4")
        vc.font   = Font(name="Calibri", size=10, color="166534")
        vc.fill   = PatternFill("solid", fgColor="F0FAF4")
        lc.border = Border(bottom=thin, top=thin, left=thin, right=thin)
        vc.border = Border(bottom=thin, top=thin, left=thin, right=thin)
        row += 1

    row += 1
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 22

    # City table
    ws3.cell(row=row, column=1, value="Engagement by City").font = Font(
        bold=True, size=12, name="Calibri", color="1A5E35")
    row += 1
    hs = header_style()
    city_display = city_stats.rename(columns={
        "city": "City", "total_providers": "Total Providers",
        "easter_providers": "Easter Providers", "engagement_pct": "Engagement %"
    })
    for c_idx, col in enumerate(city_display.columns, 1):
        cell = ws3.cell(row=row, column=c_idx, value=col)
        apply_style(cell, **hs)
        ws3.column_dimensions[get_column_letter(c_idx)].width = max(18, len(col) + 4)
    row += 1
    city_table_start = row
    for r_idx, data_row in enumerate(city_display.itertuples(index=False)):
        fill_color = ALT_ROW_BG if r_idx % 2 == 0 else "FFFFFF"
        for c_idx, val in enumerate(data_row, 1):
            cell = ws3.cell(row=row, column=c_idx, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")
        row += 1
    city_table_end = row - 1

    row += 2

    # Kyiv zone table
    ws3.cell(row=row, column=1, value="Kyiv — Engagement by Zone").font = Font(
        bold=True, size=12, name="Calibri", color="1A5E35")
    row += 1
    zone_display = zone_stats.rename(columns={
        "zone": "Zone Name", "total_providers": "Total Providers",
        "easter_providers": "Easter Providers", "engagement_pct": "Engagement %"
    })
    for c_idx, col in enumerate(zone_display.columns, 1):
        cell = ws3.cell(row=row, column=c_idx, value=col)
        apply_style(cell, **hs)
        ws3.column_dimensions[get_column_letter(c_idx)].width = max(22, len(col) + 4)
    row += 1
    for r_idx, data_row in enumerate(zone_display.itertuples(index=False)):
        fill_color = ALT_ROW_BG if r_idx % 2 == 0 else "FFFFFF"
        for c_idx, val in enumerate(data_row, 1):
            cell = ws3.cell(row=row, column=c_idx, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.border    = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")
        row += 1

    # Bar chart: city engagement
    if city_table_end > city_table_start:
        chart = BarChart()
        chart.type          = "bar"
        chart.title         = "Easter Engagement by City"
        chart.y_axis.title  = "City"
        chart.x_axis.title  = "% of Providers"
        chart.style         = 10
        chart.width         = 25
        chart.height        = max(10, (city_table_end - city_table_start + 1) * 0.6)
        data_ref = Reference(ws3, min_col=4, max_col=4,
                             min_row=city_table_start - 1, max_row=city_table_end)
        cats     = Reference(ws3, min_col=1,
                             min_row=city_table_start, max_row=city_table_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = "1DC462"
        ws3.add_chart(chart, f"F{city_table_start - 1}")

    wb.save(output_path)
    print(f"\n✅  Report saved → {output_path}")
    return output_path


# ─── JSON EXPORT FOR MAP ────────────────────────────────────────────────────────

def build_data_json(df_all: pd.DataFrame, df_easter: pd.DataFrame, output_path: str):
    """Export data.json consumed by the interactive map for live refresh."""

    # Easter partners list for map markers
    ep_cols = ["provider_id", "brand_name", "latitude", "longitude", "delivery_radius_km", "city"]
    ep_cols_present = [c for c in ep_cols if c in df_easter.columns]
    easter_list = []
    for row in df_easter[ep_cols_present].itertuples(index=False):
        r = row._asdict()
        try:
            lat = float(r.get("latitude", 0) or 0)
            lng = float(r.get("longitude", 0) or 0)
            radius = float(r.get("delivery_radius_km", 3) or 3)
            pid = int(r.get("provider_id", 0) or 0)
            name = str(r.get("brand_name", "") or "")
            city = str(r.get("city", "") or "")
            if lat and lng:
                easter_list.append([pid, name, round(lat, 6), round(lng, 6), radius, city])
        except Exception:
            pass

    # Engagement by city
    easter_ids = set(df_easter["provider_id"].dropna().astype(int).tolist())
    city_stats_list = []
    kyiv_zone_stats_list = []

    if not df_all.empty:
        df_tmp = df_all.copy()
        df_tmp["is_easter"] = df_tmp["provider_id"].astype(float).astype("Int64").isin(easter_ids)
        cs = (
            df_tmp.groupby("city")
            .agg(total=("provider_id", "count"), easter=("is_easter", "sum"))
            .reset_index()
        )
        cs["pct"] = (cs["easter"] / cs["total"] * 100).round(1)
        cs = cs.sort_values("easter", ascending=False)
        for row in cs.itertuples(index=False):
            city_stats_list.append({
                "city": str(row.city),
                "total": int(row.total),
                "easter": int(row.easter),
                "pct": float(row.pct),
            })

        # Kyiv zone breakdown
        kyiv_mask = df_tmp["city"].str.lower().str.contains("kyiv|kiev|київ", na=False)
        kyiv_df = df_tmp[kyiv_mask]
        if not kyiv_df.empty and "zone" in kyiv_df.columns:
            zs = (
                kyiv_df.groupby("zone")
                .agg(total=("provider_id", "count"), easter=("is_easter", "sum"))
                .reset_index()
            )
            zs["pct"] = (zs["easter"] / zs["total"] * 100).round(1)
            zs = zs.sort_values("easter", ascending=False)
            for row in zs.itertuples(index=False):
                kyiv_zone_stats_list.append({
                    "zone": str(row.zone),
                    "total": int(row.total),
                    "easter": int(row.easter),
                    "pct": float(row.pct),
                })
    else:
        # Fallback: only easter-side city counts
        cs = df_easter.groupby("city").agg(easter=("provider_id", "count")).reset_index()
        for row in cs.sort_values("easter", ascending=False).itertuples(index=False):
            city_stats_list.append({"city": str(row.city), "total": int(row.easter),
                                    "easter": int(row.easter), "pct": 100.0})

    # Normalize city names to match zone-polygon names
    CITY_NORM = {
        "Zaporizhia": "Zaporizhzhia",
        "Zaporizhzhya": "Zaporizhzhia",
    }
    UA_CITIES = {
        "Bila Tserkva", "Boryspil", "Brovary", "Bucha", "Cherkasy",
        "Chernihiv", "Chernivtsi", "Dnipro", "Drohobych", "Irpin",
        "Ivano-Frankivsk", "Kamianets-Podilskyi", "Kamianske", "Kharkiv",
        "Khmelnytskyi", "Kovel", "Kremenchuk", "Kropyvnytskyi", "Kryvyi Rih",
        "Kyiv", "Lutsk", "Lviv", "Mukachevo", "Mykolaiv", "Odesa",
        "Oleksandriia", "Pavlohrad", "Poltava", "Rivne", "Sumy",
        "Ternopil", "Uzhhorod", "Vinnytsia", "Vyshhorod",
        "Zaporizhzhia", "Zhytomyr",
    }
    for r in city_stats_list:
        r["city"] = CITY_NORM.get(r["city"], r["city"])
    for p in easter_list:
        p[5] = CITY_NORM.get(p[5], p[5])
    city_stats_list = [r for r in city_stats_list if r["city"] in UA_CITIES]

    payload = {
        "date": REPORT_DATE,
        "total_providers": len(df_all) if not df_all.empty else 0,
        "total_easter": len(df_easter),
        "easter_partners": easter_list,
        "city_stats": city_stats_list,
        "kyiv_zones": kyiv_zone_stats_list,
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    print(f"✅  data.json saved → {output_path}  ({len(easter_list):,} partners, {len(city_stats_list)} cities)")
    return output_path


# ─── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    print(f"=== Bolt Food Partner Report  [{REPORT_DATE}] ===\n")

    df_easter = fetch_easter_providers()

    try:
        df_all = fetch_all_providers()
    except Exception as exc:
        print(f"  ! Could not fetch all providers from Databricks: {exc}")
        print("  → Continuing with Easter data only")
        df_all = pd.DataFrame()

    base_dir = os.path.dirname(__file__)
    out_path = os.path.join(base_dir, OUTPUT_FILE)
    build_report(df_all, df_easter, out_path)

    json_path = os.path.join(base_dir, "data.json")
    build_data_json(df_all, df_easter, json_path)

    return out_path


if __name__ == "__main__":
    main()
